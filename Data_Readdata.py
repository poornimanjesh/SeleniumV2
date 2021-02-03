import openpyxl
path=r"C:\users\manju\Documents\Poornima\excel\openpy.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
rows = sheet.max_row
col = sheet.max_column
print(rows)
print(col)
for r in range(1,rows+1):
    for c in range(1,col+1):
        print(sheet.cell(row=r,column=c).value,end="  ")
    print()